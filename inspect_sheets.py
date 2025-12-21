import pandas as pd
from openpyxl import load_workbook

files = ["MEPL.xlsx", "MLPL.xlsx"]

for file in files:
    print(f"--- Inspecting Sheets in {file} ---")
    try:
        wb = load_workbook(file, read_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet in wb.sheetnames:
            print(f"Sheet: {sheet}")
            df = pd.read_excel(file, sheet_name=sheet, skiprows=1)
            if 'Buying legal entity' in df.columns:
                 print(f"Unique values in 'Buying legal entity': {df['Buying legal entity'].unique()}")
            else:
                print("Column 'Buying legal entity' not found.")
            print("-" * 10)
            
    except Exception as e:
        print(f"Error reading {file}: {e}")
